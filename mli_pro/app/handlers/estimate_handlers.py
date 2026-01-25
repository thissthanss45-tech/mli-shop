import os
import openpyxl
from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile

# Добавили get_user_role чтобы проверять права
from app.db import get_my_objects, get_sectors_by_object, save_estimate_row, get_estimate_v2, get_user_role
from app.keyboards import get_work_objects_kb, get_sectors_kb

router = Router()

class EstimateState(StatesGroup):
    waiting_for_obj = State()
    waiting_for_sector = State()
    waiting_for_file = State()

# --- МЕНЮ СМЕТ ---
@router.message(F.text == "📋 Сметы")
async def estimate_menu(message: types.Message):
    buttons = [
        [InlineKeyboardButton(text="📥 Загрузить Excel", callback_data="upload_est")],
        [InlineKeyboardButton(text="👁 Посмотреть смету", callback_data="view_est")],
        [InlineKeyboardButton(text="📄 Скачать шаблон", callback_data="get_template")]
    ]
    await message.answer("🛠 Управление сметами:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))

# --- ПРОСМОТР (С ЦЕНАМИ ДЛЯ АДМИНА) ---
@router.callback_query(F.data == "view_est")
async def view_est_start(callback: types.CallbackQuery):
    objects = get_my_objects(callback.from_user.id)
    if not objects:
        await callback.message.answer("Нет объектов.")
        return
    buttons = [[InlineKeyboardButton(text=f"👁 {name}", callback_data=f"look_obj_{obj_id}")] for obj_id, name, s in objects]
    await callback.message.answer("Выберите объект:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@router.callback_query(F.data.startswith("look_obj_"))
async def view_est_show(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    obj_id = int(callback.data.split("_")[2])
    
    # 1. Получаем роль
    role = get_user_role(user_id)
    
    # 2. Получаем данные (Название, Ед, Кол-во, Цена)
    rows = get_estimate_v2(obj_id)
    
    if not rows:
        await callback.message.answer("Пусто. Загрузите файл.")
        return

    text = f"📋 **СМЕТА ОБЪЕКТА (Вы видите её как: {role.upper()})**\n\n"
    
    total_budget = 0
    
    for r in rows:
        # r[0]=Name, r[1]=Unit, r[2]=Qty, r[3]=Price
        name = r[0]
        unit = r[1]
        qty = r[2]
        price = r[3]
        
        # Базовая строка (видят все)
        line = f"🔹 {name}: {qty} {unit}"
        
        # ДОБАВКА: Если Админ - показываем цену
        if role == 'admin':
            sum_pos = qty * price
            line += f" | 💰 {price} р. = {sum_pos:,.0f} р."
            total_budget += sum_pos
            
        text += line + "\n"
    
    # ИТОГО (Только Админ)
    if role == 'admin':
        text += f"\n💵 **ИТОГО БЮДЖЕТ: {total_budget:,.0f} руб.**"
    
    await callback.message.answer(text[:4000]) # Ограничение Телеграма на длину
    await callback.answer()

# --- ЗАГРУЗКА (ЛОГИКА НЕ МЕНЯЛАСЬ) ---
@router.callback_query(F.data == "upload_est")
async def upload_start(callback: types.CallbackQuery, state: FSMContext):
    objects = get_my_objects(callback.from_user.id)
    if not objects:
        await callback.message.answer("❌ Нет объектов.")
        return
    await callback.message.answer("1️⃣ Выберите объект:", reply_markup=get_work_objects_kb(objects))
    await state.set_state(EstimateState.waiting_for_obj)
    await callback.answer()

@router.callback_query(EstimateState.waiting_for_obj, F.data.startswith("work_obj_"))
async def upload_sector_step(callback: types.CallbackQuery, state: FSMContext):
    obj_id = int(callback.data.split("_")[2])
    await state.update_data(current_obj_id=obj_id)
    sectors = get_sectors_by_object(obj_id)
    if sectors:
        await callback.message.answer("2️⃣ Выберите сектор:", reply_markup=get_sectors_kb(sectors))
        await state.set_state(EstimateState.waiting_for_sector)
    else:
        await state.update_data(current_sector_id=None)
        await ask_for_file(callback, state)
    await callback.answer()

@router.callback_query(EstimateState.waiting_for_sector, F.data.startswith("work_sect_"))
async def upload_file_step(callback: types.CallbackQuery, state: FSMContext):
    data_str = callback.data.split("_")[2]
    sector_id = int(data_str) if data_str != "none" else None
    await state.update_data(current_sector_id=sector_id)
    await ask_for_file(callback, state)

async def ask_for_file(callback, state):
    await callback.message.answer("3️⃣ Пришлите Excel (.xlsx)")
    await state.set_state(EstimateState.waiting_for_file)

@router.message(EstimateState.waiting_for_file, F.document)
async def process_file(message: types.Message, state: FSMContext):
    document = message.document
    file_path = f"temp_{document.file_name}"
    await message.bot.download(document, destination=file_path)
    data = await state.get_data()
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            price = 0
            if len(row) > 3 and row[3]:
                try:
                    price = float(row[3])
                except:
                    price = 0
            save_estimate_row(data['current_obj_id'], data['current_sector_id'], str(row[0]), str(row[1] or "шт"), float(row[2] or 0), price)
            count += 1
        await message.answer(f"✅ Загружено {count} строк.")
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")
    finally:
        if os.path.exists(file_path): os.remove(file_path)
        await state.clear()

@router.callback_query(F.data == "get_template")
async def get_template(callback: types.CallbackQuery):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Название", "Ед", "Кол-во", "Цена"])
    ws.append(["Тест", "шт", "1", "1000"])
    wb.save("temp.xlsx")
    await callback.message.answer_document(FSInputFile("temp.xlsx"))
    os.remove("temp.xlsx")
    await callback.answer()