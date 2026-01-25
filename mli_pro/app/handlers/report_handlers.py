import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from aiogram import Router, F, types
from aiogram.types import FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton

from app.db import get_my_objects, get_ks2_data, get_user_role

router = Router()

# 1. МЕНЮ ОТЧЕТОВ
# 🔥 ИСПРАВЛЕНО: Теперь ловим короткое название, как в клавиатуре
@router.message(F.text == "📊 Отчеты")
@router.message(F.text == "📊 Отчеты и КС-2") # На всякий случай ловим и старое
async def reports_menu(message: types.Message):
    user_id = message.from_user.id
    objects = get_my_objects(user_id)
    
    if not objects:
        await message.answer("Нет доступных объектов.")
        return

    # Клавиатура выбора объекта
    buttons = []
    for obj_id, name, status in objects:
        buttons.append([InlineKeyboardButton(text=f"📊 {name}", callback_data=f"gen_report_{obj_id}")])
        
    await message.answer("По какому объекту сформировать Акт (КС-2)?", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))

# 2. ГЕНЕРАЦИЯ EXCEL
@router.callback_query(F.data.startswith("gen_report_"))
async def generate_report(callback: types.CallbackQuery):
    obj_id = int(callback.data.split("_")[2])
    user_id = callback.from_user.id
    role = get_user_role(user_id)
    
    # Получаем данные из БД
    data = get_ks2_data(obj_id)
    
    if not data:
        await callback.message.answer("❌ Данных нет. Смета пуста или работы не велись.")
        await callback.answer()
        return

    await callback.message.answer("⏳ Формирую документ... Подождите.")

    # --- СОЗДАНИЕ EXCEL ---
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Акт КС-2"
        
        # Заголовки
        headers = ["Наименование работ", "Ед.изм", "Цена (Руб)", "План (Кол-во)", "ФАКТ (Кол-во)", "Остаток", "Сумма Выполнения (Руб)"]
        ws.append(headers)
        
        # Стили
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Красим шапку
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        total_money = 0
        
        for row in data:
            # row: (name, unit, price, plan, fact)
            name, unit, price, plan, fact = row
            
            # Если юзер НЕ Админ — скрываем цены (ставим 0)
            if role != 'admin':
                price = 0
                sum_money = 0
            else:
                sum_money = fact * price
                
            remain = plan - fact
            total_money += sum_money
            
            # Записываем строку
            ws.append([name, unit, price, plan, fact, remain, sum_money])

        # Итоговая строка (Только для админа)
        if role == 'admin':
            last_row = len(data) + 2
            ws.cell(row=last_row, column=6, value="ИТОГО:").font = bold_font
            ws.cell(row=last_row, column=7, value=total_money).font = bold_font

        # Автоширина колонок
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['G'].width = 20
        
        # Сохраняем во временный файл
        filename = f"KS2_Object_{obj_id}.xlsx"
        wb.save(filename)
        
        # Отправляем файл пользователю
        await callback.message.answer_document(
            FSInputFile(filename),
            caption=f"📊 **Акт выполнения (КС-2)**\n{'Сумма: ' + str(total_money) + ' руб.' if role == 'admin' else ''}"
        )
        
        # Удаляем файл с сервера
        os.remove(filename)
    
    except Exception as e:
        await callback.message.answer(f"❌ Ошибка генерации: {e}")
        
    await callback.answer()