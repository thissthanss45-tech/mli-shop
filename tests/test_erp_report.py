from __future__ import annotations

import os
import tempfile
import unittest
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

os.environ.setdefault("BOT_TOKEN", "test-token")
os.environ.setdefault("OWNER_ID", "1")
os.environ.setdefault("WEB_ADMIN_KEY", "test-admin-key")
os.environ.setdefault("CORS_ORIGINS", "https://example.com")
os.environ.setdefault("DB_URL", "sqlite+aiosqlite:///test_erp_report_placeholder.db")

from database.db_manager import Base  # noqa: E402
from models import Brand, Category, Product, StockMovement, Tenant  # noqa: E402
from utils.erp_report import build_erp_report_xlsx  # noqa: E402


async def _create_schema(engine) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


class ErpReportTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self._tmp_dir = tempfile.TemporaryDirectory()
        db_path = os.path.join(self._tmp_dir.name, "erp_report.db")
        self.engine = create_async_engine(
            f"sqlite+aiosqlite:///{db_path}",
            connect_args={"check_same_thread": False},
            poolclass=NullPool,
        )
        self.maker = async_sessionmaker(self.engine, expire_on_commit=False, class_=AsyncSession)
        await _create_schema(self.engine)

    async def asyncTearDown(self) -> None:
        await self.engine.dispose()
        self._tmp_dir.cleanup()

    async def test_movement_sheet_uses_owner_journal_format(self) -> None:
        async with self.maker() as session:
            tenant = Tenant(slug="flowers-boutique", title="Flowers Boutique", status="active")
            session.add(tenant)
            await session.flush()

            brand = Brand(name="Rose Studio", tenant_id=tenant.id)
            category = Category(name="Розы", tenant_id=tenant.id)
            session.add_all([brand, category])
            await session.flush()

            product = Product(
                tenant_id=tenant.id,
                sku="ROS-0001",
                title="Роза алая",
                purchase_price=100,
                sale_price=200,
                category_id=category.id,
                brand_id=brand.id,
            )
            session.add(product)
            await session.flush()

            movement = StockMovement(
                tenant_id=tenant.id,
                product_id=product.id,
                order_id=15,
                size="шт",
                quantity=2,
                stock_before=7,
                stock_after=5,
                direction="out",
                operation_type="sale",
                note="",
                created_at=datetime.utcnow(),
            )
            session.add(movement)
            await session.commit()

            payload = await build_erp_report_xlsx(
                session,
                datetime.utcnow() - timedelta(days=1),
                datetime.utcnow() + timedelta(days=1),
            )

        workbook = load_workbook(BytesIO(payload))
        sheet = workbook["Движение"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                "Когда",
                "Товар",
                "Что произошло",
                "Сколько",
                "Комментарий",
            ],
        )

        first_row = [sheet.cell(row=2, column=index).value for index in range(1, 6)]
        self.assertEqual(first_row[1], "Rose Studio / Роза алая")
        self.assertEqual(first_row[2], "Продали товар")
        self.assertEqual(first_row[3], "-2")
        self.assertEqual(first_row[4], "Заказ №15")


if __name__ == "__main__":
    unittest.main(verbosity=2)