from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select, func, case
from sqlalchemy.ext.asyncio import AsyncSession

from models import Brand, Order, OrderItem, OrderStatus, Product, ProductStock, StockMovement


def _movement_operation_label(operation_type: str) -> str:
    raw_value = getattr(operation_type, "value", operation_type)
    mapping = {
        "sale": "Продали товар",
        "manual_add": "Добавили на склад",
        "manual_write_off": "Убрали со склада",
        "return": "Вернули на склад",
        "correction": "Исправили остаток",
    }
    normalized = (raw_value or "").strip().lower()
    return mapping.get(normalized, raw_value or "")


def _movement_delta_value(direction: str, quantity: int) -> str:
    raw_value = getattr(direction, "value", direction)
    sign = "+" if raw_value == "in" else "-"
    return f"{sign}{int(quantity)}"


def _movement_comment_text(note: str | None, operation_type: str, order_id: int | None) -> str:
    cleaned_note = (note or "").strip()
    raw_value = getattr(operation_type, "value", operation_type)
    if cleaned_note:
        return cleaned_note
    if order_id and raw_value == "sale":
        return f"Заказ №{order_id}"
    if order_id and raw_value == "return":
        return f"Возврат по заказу №{order_id}"
    if raw_value == "manual_add":
        return "Товар пришёл в магазин"
    if raw_value == "manual_write_off":
        return "Товар списали вручную"
    if raw_value == "correction":
        return "Остаток исправлен вручную"
    return "Без комментария"


def _movement_product_label(brand_name: str | None, product_title: str | None) -> str:
    title = (product_title or "").strip()
    brand = (brand_name or "").strip()
    if brand and title:
        return f"{brand} / {title}"
    if title:
        return title
    if brand:
        return brand
    return "Товар не указан"


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[column].width = min(max(max_len + 2, 12), 60)


def _style_sheet_header(ws) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


async def build_erp_report_xlsx(
    session: AsyncSession,
    start_date: datetime,
    end_date: datetime,
) -> bytes:
    sales_stmt = (
        select(
            Order.created_at,
            Order.id.label("order_id"),
            Product.sku,
            Brand.name.label("brand"),
            Product.title,
            OrderItem.size,
            OrderItem.quantity,
            OrderItem.sale_price,
            Product.purchase_price,
        )
        .join(OrderItem, Order.id == OrderItem.order_id)
        .join(Product, Product.id == OrderItem.product_id)
        .join(Brand, Brand.id == Product.brand_id)
        .where(
            Order.status == OrderStatus.COMPLETED.value,
            Order.created_at.between(start_date, end_date),
        )
        .order_by(Order.created_at.asc(), Order.id.asc())
    )
    sales_rows = (await session.execute(sales_stmt)).all()

    sales_data = []
    for row in sales_rows:
        sale_total = float(row.sale_price) * int(row.quantity)
        purchase_total = float(row.purchase_price) * int(row.quantity)
        profit = sale_total - purchase_total
        sales_data.append(
            {
                "Дата": row.created_at.strftime("%d.%m.%Y %H:%M"),
                "Заказ": row.order_id,
                "SKU": row.sku,
                "Бренд": row.brand,
                "Товар": row.title,
                "Размер": row.size,
                "Кол-во": int(row.quantity),
                "Закупка/ед": float(row.purchase_price),
                "Продажа/ед": float(row.sale_price),
                "Выручка": sale_total,
                "Прибыль": profit,
            }
        )

    stock_stmt = (
        select(
            Product.id.label("product_id"),
            Product.sku,
            Brand.name.label("brand"),
            Product.title,
            ProductStock.size,
            ProductStock.quantity,
            Product.purchase_price,
            Product.sale_price,
        )
        .join(ProductStock, Product.id == ProductStock.product_id)
        .join(Brand, Brand.id == Product.brand_id)
        .order_by(Brand.name.asc(), Product.title.asc(), ProductStock.size.asc())
    )
    stock_rows = (await session.execute(stock_stmt)).all()

    incoming_summary_stmt = (
        select(
            StockMovement.product_id,
            StockMovement.size,
            func.coalesce(
                func.sum(
                    case((StockMovement.direction == "in", StockMovement.quantity), else_=0)
                ),
                0,
            ).label("incoming_qty"),
        )
        .where(StockMovement.created_at.between(start_date, end_date))
        .group_by(StockMovement.product_id, StockMovement.size)
    )
    incoming_summary_rows = (await session.execute(incoming_summary_stmt)).all()
    incoming_map = {
        (int(row.product_id), str(row.size)): int(row.incoming_qty or 0)
        for row in incoming_summary_rows
    }

    sales_outgoing_stmt = (
        select(
            OrderItem.product_id,
            OrderItem.size,
            func.coalesce(func.sum(OrderItem.quantity), 0).label("sold_qty"),
        )
        .join(Order, Order.id == OrderItem.order_id)
        .where(
            Order.status == OrderStatus.COMPLETED.value,
            Order.created_at.between(start_date, end_date),
        )
        .group_by(OrderItem.product_id, OrderItem.size)
    )
    sales_outgoing_rows = (await session.execute(sales_outgoing_stmt)).all()
    sold_map = {
        (int(row.product_id), str(row.size)): int(row.sold_qty or 0)
        for row in sales_outgoing_rows
    }

    stock_data = []
    for row in stock_rows:
        qty = int(row.quantity)
        purchase = float(row.purchase_price)
        sale = float(row.sale_price)
        key = (int(row.product_id), str(row.size))
        incoming_qty = incoming_map.get(key, 0)
        outgoing_qty = sold_map.get(key, 0)
        if incoming_qty == 0 and key not in incoming_map and qty > 0:
            incoming_qty = qty
        stock_data.append(
            {
                "SKU": row.sku,
                "Бренд": row.brand,
                "Товар": row.title,
                "Размер": row.size,
                "Приход": incoming_qty,
                "Расход": outgoing_qty,
                "Остаток": qty,
                "Закупка/ед": purchase,
                "Продажа/ед": sale,
                "Заморожено (закупка)": qty * purchase,
                "Потенц. выручка": qty * sale,
            }
        )

    movement_stmt = (
        select(
            StockMovement.created_at,
            StockMovement.direction,
            StockMovement.operation_type,
            StockMovement.order_id,
            StockMovement.size,
            StockMovement.quantity,
            StockMovement.stock_before,
            StockMovement.stock_after,
            StockMovement.note,
            Product.sku,
            Product.title,
            Brand.name.label("brand"),
        )
        .join(Product, Product.id == StockMovement.product_id)
        .join(Brand, Brand.id == Product.brand_id)
        .where(StockMovement.created_at.between(start_date, end_date))
        .order_by(StockMovement.created_at.asc(), StockMovement.id.asc())
    )
    movement_rows = (await session.execute(movement_stmt)).all()

    movement_data = []
    for row in movement_rows:
        movement_data.append(
            {
                "Когда": row.created_at.strftime("%d.%m.%Y %H:%M"),
                "Товар": _movement_product_label(row.brand, row.title),
                "Что произошло": _movement_operation_label(row.operation_type),
                "Сколько": _movement_delta_value(row.direction, row.quantity),
                "Комментарий": _movement_comment_text(row.note, row.operation_type, row.order_id),
            }
        )

    if not movement_data and sales_data:
        for sale in sales_data:
            movement_data.append(
                {
                    "Когда": sale["Дата"],
                    "Товар": _movement_product_label(sale["Бренд"], sale["Товар"]),
                    "Что произошло": "Продали товар",
                    "Сколько": f"-{sale['Кол-во']}",
                    "Комментарий": f"Заказ №{sale['Заказ']}",
                }
            )

    if sales_data:
        total_revenue = sum(float(item["Выручка"]) for item in sales_data)
        total_profit = sum(float(item["Прибыль"]) for item in sales_data)
        sales_data.append(
            {
                "Дата": "",
                "Заказ": "",
                "SKU": "",
                "Бренд": "",
                "Товар": "ИТОГО (₽)",
                "Размер": "",
                "Кол-во": "",
                "Закупка/ед": "",
                "Продажа/ед": "",
                "Выручка": round(total_revenue, 2),
                "Прибыль": round(total_profit, 2),
            }
        )

    if stock_data:
        total_incoming = sum(int(item["Приход"]) for item in stock_data)
        total_outgoing = sum(int(item["Расход"]) for item in stock_data)
        total_frozen = sum(float(item["Заморожено (закупка)"]) for item in stock_data)
        total_potential = sum(float(item["Потенц. выручка"]) for item in stock_data)
        stock_data.append(
            {
                "SKU": "",
                "Бренд": "",
                "Товар": "ИТОГО (₽)",
                "Размер": "",
                "Приход": total_incoming,
                "Расход": total_outgoing,
                "Остаток": "",
                "Закупка/ед": "",
                "Продажа/ед": "",
                "Заморожено (закупка)": round(total_frozen, 2),
                "Потенц. выручка": round(total_potential, 2),
            }
        )

    sales_df = pd.DataFrame(sales_data)
    stock_df = pd.DataFrame(stock_data)
    movement_df = pd.DataFrame(movement_data)

    if sales_df.empty:
        sales_df = pd.DataFrame(columns=[
            "Дата", "Заказ", "SKU", "Бренд", "Товар", "Размер", "Кол-во", "Закупка/ед", "Продажа/ед", "Выручка", "Прибыль"
        ])
    if stock_df.empty:
        stock_df = pd.DataFrame(columns=[
            "SKU", "Бренд", "Товар", "Размер", "Приход", "Расход", "Остаток", "Закупка/ед", "Продажа/ед", "Заморожено (закупка)", "Потенц. выручка"
        ])
    if movement_df.empty:
        movement_df = pd.DataFrame(columns=[
            "Когда", "Товар", "Что произошло", "Сколько", "Комментарий"
        ])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sales_df.to_excel(writer, index=False, sheet_name="Продажи")
        stock_df.to_excel(writer, index=False, sheet_name="Склад")
        movement_df.to_excel(writer, index=False, sheet_name="Движение")

        wb = writer.book
        ws_sales = wb["Продажи"]
        ws_stock = wb["Склад"]
        ws_mov = wb["Движение"]

        for ws in (ws_sales, ws_stock, ws_mov):
            _style_sheet_header(ws)
            _autosize_columns(ws)

        ws_mov.auto_filter.ref = ws_mov.dimensions

        if ws_mov.max_row > 1:
            in_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            out_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            for row_idx in range(2, ws_mov.max_row + 1):
                delta_value = str(ws_mov.cell(row=row_idx, column=4).value or "")
                fill = in_fill if delta_value.startswith("+") else out_fill
                for col_idx in range(1, ws_mov.max_column + 1):
                    ws_mov.cell(row=row_idx, column=col_idx).fill = fill

    output.seek(0)
    return output.getvalue()
